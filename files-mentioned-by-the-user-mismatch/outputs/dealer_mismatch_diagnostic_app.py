import math
import re
import sys
from datetime import datetime
from dataclasses import dataclass
from pathlib import Path
from tkinter import BOTH, END, LEFT, RIGHT, TOP, X, Y, filedialog, messagebox
import tkinter as tk
from tkinter import simpledialog, ttk

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


APP_TITLE = "经销商库存对账 Mismatch 智能诊断系统"
EXPECTED_SHEETS = [
    "Summary",
    "Only_in_List",
    "Only_in_SAP",
    "Mismatch_Detail",
    "Stock_Base",
    "Actual feedback",
]

KEY_COLUMNS_AFTER_CHASSIS = [
    "Rule_ID",
    "Diagnostic category",
    "Data-chain judgement",
    "Process issues identified",
    "How to optimise",
    "Process issue category",
    "Root cause hypothesis",
    "Control gap",
    "Recommended control",
    "Required evidence",
    "Next action",
    "Preventive rule",
    "Linked example cases",
    "Priority",
    "Confidence",
    "Action owner",
    "Stock_Type",
    "Lifecycle_Stage",
    "Mismatch_Type",
]

DISPLAY_HIDDEN_COLUMNS = {"Actual situation after check"}


RULE_COLORS = {
    "High": "F8D7DA",
    "Medium": "FFF3CD",
    "Low": "DDEEDB",
    "Header": "DCE7F1",
}


CHINESE_COLUMN_NAMES = {
    "Rule_ID": "规则编号",
    "Diagnostic category": "诊断类别",
    "Data-chain judgement": "数据链判断",
    "Evidence-based reason": "证据原因",
    "Process issues identified": "识别出的流程问题",
    "How to optimise": "优化建议",
    "Process issue category": "流程问题分类",
    "Root cause hypothesis": "根因假设",
    "Control gap": "控制缺口",
    "Recommended control": "建议控制措施",
    "Required evidence": "所需证据",
    "Next action": "下一步动作",
    "Preventive rule": "预防规则",
    "Linked example cases": "关联示例案例",
    "Priority": "优先级",
    "Confidence": "置信度",
    "Action owner": "行动负责人",
    "Stock_Type": "库存类型",
    "Lifecycle_Stage": "生命周期阶段",
    "Mismatch_Type": "不匹配类型",
    "Normalized_Chassis": "标准化底盘号",
    "Original_Chassis": "原始底盘号/VIN",
    "Actual situation after check": "核查后的实际情况",
    "Metric": "指标",
    "Value": "数值",
    "Sheet": "工作表",
    "Rows": "行数",
    "Columns": "列数",
    "Empty_Rate_%": "空值率_%",
    "Status": "状态",
}

CHINESE_VALUE_TRANSLATIONS = {
    "High": "高",
    "Medium": "中",
    "Low": "低",
    "Data": "数据团队",
    "Dealer": "经销商",
    "SAP support": "SAP支持",
    "Finance": "财务",
    "Sales admin": "销售行政",
    "Normal new stock": "普通新车库存",
    "Unclassified stock": "未分类库存",
    "Used stock / placeholder": "二手车/占位库存",
    "Demo stock": "试驾/展示车库存",
    "Category 1": "类别1",
    "Category 2": "类别2",
    "False Positive - Naming": "假阳性-命名/录入差异",
    "SAP No Evidence": "SAP无证据",
    "Process Break": "流程断点",
    "Likely Data Timing Delay": "可能是数据时点延迟",
    "Completed / Left Stock": "已完成/已离库",
    "Needs Manual Review": "需要人工复核",
    "Cancelled Sale / Refund Not Closed": "取消销售/退款未闭环",
    "PGI Closure Violation": "PGI闭环违规",
    "Reservation Resale Risk": "预订后转售风险",
    "Demo / Used Special Handling": "试驾/二手特殊处理",
    "Upstream Shipped / Dealer Not Updated": "上游已发运/经销商长期未更新",
    "Total mismatch records": "Mismatch总记录数",
    "No strong lifecycle evidence found.": "未发现强生命周期证据。",
    "OK": "正常",
    "Extra": "额外",
    "Missing": "缺失",
}

CHINESE_PHRASE_TRANSLATIONS = {
    "Chassis format or one-character input difference created a false mismatch.": "底盘号格式或单字符录入差异造成了假 mismatch。",
    "Apply Normalized_Chassis matching before generating mismatch lists.": "生成 mismatch 清单前先执行标准化底盘号匹配。",
    "The vehicle appears on the dealer list, but SAP lifecycle evidence is not available.": "车辆出现在经销商清单中，但SAP生命周期证据不可用。",
    "SAP master/lifecycle data may be missing, delayed or entered under a different identifier.": "SAP主数据/生命周期数据可能缺失、延迟，或录入在不同识别号下。",
    "Check source entry, SAP creation status and Chassis normalization before treating it as true stock.": "先检查源数据录入、SAP创建状态和底盘号标准化，再判断为真实库存异常。",
    "Downstream PGI exists but the expected dealer SO evidence is missing.": "存在下游PGI，但缺少预期的经销商SO证据。",
    "A downstream document exists without the upstream document normally required by the sales chain.": "存在下游单据，但缺少销售链通常要求的上游单据。",
    "Audit document lineage and prevent downstream PGI records from being accepted without a linked SO.": "审计单据链路，防止无关联SO的下游PGI被接受。",
    "Dealer PGI date appears earlier than factory PGI date.": "经销商PGI日期早于工厂PGI日期。",
    "The expected factory-to-dealer movement sequence may be reversed or incorrectly dated.": "工厂到经销商的预期移动顺序可能被反向记录或日期录入错误。",
    "Audit movement timestamps and correct the document sequence if dates were posted out of order.": "审计移动时间戳；如日期过账顺序错误，则纠正单据顺序。",
    "Dealer PGI is the latest visible movement and no dealer invoice is available.": "经销商PGI是最新可见移动，且没有经销商发票证据。",
    "PGI has been completed and not reversed, but invoice evidence is missing.": "PGI已完成且未冲销，但缺少发票证据。",
    "Create an automatic PGI-without-invoice exception. Confirm whether the unit has left stock, invoice is delayed, or PGI should be reversed.": "建立自动的“PGI无发票”异常；确认车辆是否已离库、发票是否延迟，或PGI是否应冲销。",
    "The data chain shows stock movement but no invoice closure.": "数据链显示库存移动，但没有发票闭环。",
    "PGI can be posted without invoice follow-up or reversal tracking.": "PGI可能在没有发票跟进或冲销跟踪的情况下被过账。",
    "Daily exception report for PGI with no invoice beyond allowed SLA; require owner, ageing and closure reason.": "对超过SLA仍无发票的PGI生成每日异常报表，并要求负责人、账龄和关闭原因。",
    "Invoice exists but Bill-to evidence is missing.": "存在发票，但缺少Bill-to证据。",
    "Billing document exists without a clear billing party.": "存在开票单据，但没有清晰的开票方。",
    "Make Bill-to mandatory for invoice-linked reconciliation and block closure until the customer account is mapped.": "在发票相关对账中将Bill-to设为必填；客户账户映射前不允许关闭。",
    "Sales order exists but PO/GR evidence is missing or zero.": "存在销售订单，但PO/GR证据缺失或为0。",
    "Procurement evidence is absent even though sales order evidence exists.": "虽然有销售订单证据，但缺少采购证据。",
    "Check whether this is internal transfer, delayed PO/GR capture or missing procurement linkage.": "检查是否为内部调拨、PO/GR采集延迟，或采购链路缺失。",
    "Reverse PGI flag and Last_Movement_Is_PGI both appear active.": "Reverse PGI标记和Last_Movement_Is_PGI同时为有效。",
    "Reversal status conflicts with the last movement status.": "冲销状态与最新移动状态冲突。",
    "Reconcile movement history and ensure reversal status updates the latest movement indicator.": "核对移动历史，并确保冲销状态同步更新最新移动标记。",
    "Dealer and factory material numbers do not match.": "经销商与工厂物料号不一致。",
    "Material mismatch can point to order creation or vehicle assignment errors.": "物料不一致可能指向订单创建或车辆分配错误。",
    "Validate material mapping between dealer SO and factory SO before closing the reconciliation.": "关闭对账前验证经销商SO与工厂SO之间的物料映射。",
    "Lifecycle evidence suggests the unit may have completed or left the expected stock state.": "生命周期证据显示车辆可能已完成流程或离开预期库存状态。",
    "List and SAP stock status may be out of sync after invoice, handover or PGI movement.": "发票、交车或PGI移动后，清单与SAP库存状态可能不同步。",
    "Confirm handover/stock location and remove or refresh stale dealer list records.": "确认交车/库存位置，并移除或刷新过期的经销商清单记录。",
    "SAP lifecycle evidence exists, but current List/SAP timing does not fully align.": "存在SAP生命周期证据，但当前清单/SAP时点未完全一致。",
    "Data refresh timing or status propagation may be behind the physical stock situation.": "数据刷新或状态传递可能滞后于实物库存状态。",
    "Wait for refresh cycle or check the list update mechanism for this dealer.": "等待刷新周期，或检查该经销商的清单更新机制。",
    "The available evidence is not enough to make a confident final judgement.": "现有证据不足以做出高置信度最终判断。",
    "Key lifecycle fields are missing or not mapped from the source workbook.": "关键生命周期字段缺失，或未从源工作簿映射出来。",
    "Collect SO, PGI, reverse PGI, invoice, PO/GR and actual feedback, then rerun diagnosis.": "补充SO、PGI、Reverse PGI、发票、PO/GR和实际反馈后重新诊断。",
    "Factory-side PGI was reversed while the dealer list still shows the unit.": "工厂侧PGI已冲销，但经销商清单仍显示该车辆。",
    "Factory PGI reversal suggests the vehicle was returned or rolled back, but dealer list/SAP stock state is not aligned.": "工厂PGI冲销表明车辆可能退回或回滚，但经销商清单/SAP库存状态未对齐。",
    "Add a reversal-close checklist: confirm reverse PGI, physical stock location, dealer list refresh and whether the dealer-side PGI/invoice also needs correction.": "增加冲销关闭清单：确认Reverse PGI、实物库存位置、经销商清单刷新，以及经销商侧PGI/发票是否也需要更正。",
    "Upstream SAP evidence indicates the vehicle was shipped long ago, but the dealer/list side still has no matching update.": "上游SAP证据显示车辆很早已发运，但经销商/清单侧仍没有匹配更新。",
    "A vehicle with upstream PGI/shipment evidence has remained absent from dealer-side updates far beyond the normal data refresh window.": "已有上游PGI/发运证据的车辆长期未出现在经销商侧更新中，远超正常数据刷新窗口。",
    "Create an ageing control for upstream-shipped vehicles that are still missing from dealer/list updates. Any PGI older than 730 days without dealer confirmation should be escalated as a process-control exception, not treated as ordinary timing delay.": "为已上游发运但经销商/清单仍未更新的车辆建立账龄控制。PGI超过730天且无经销商确认的记录，应升级为流程控制异常，而不是普通时点延迟。",
    "The upstream system shows the vehicle was shipped, but dealer-side ownership/list status was never confirmed or refreshed after a long ageing period.": "上游系统显示车辆已发运，但经过长期账龄后，经销商侧归属/清单状态仍未确认或刷新。",
    "Long-aged upstream shipment exceptions can remain in a generic timing bucket without a mandatory dealer confirmation SLA.": "长期账龄的上游发运异常可能停留在普通时点延迟分类中，缺少强制经销商确认SLA。",
    "Run a monthly ageing report for Only-in-SAP records with PGI/shipment dates older than 730 days and no dealer list update. Require dealer confirmation, physical location proof and list refresh evidence.": "每月生成账龄报表，筛查PGI/发运日期超过730天且无经销商清单更新的Only-in-SAP记录；要求提供经销商确认、实物位置证明和清单刷新证据。",
    "PGI/shipment document and date, dealer list refresh history, dealer physical confirmation, ownership/transfer record and final stock status.": "PGI/发运单据及日期、经销商清单刷新历史、经销商实物确认、归属/转移记录和最终库存状态。",
    "Dealer confirms whether the vehicle was received/sold/transferred; Data team checks list refresh history; SAP support validates upstream movement and ownership chain.": "经销商确认车辆是否已接收/销售/转移；数据团队检查清单刷新历史；SAP支持验证上游移动和归属链路。",
    "If upstream PGI/shipment evidence is older than 730 days and dealer/list update is missing, auto-escalate as a high-confidence ageing exception.": "如果上游PGI/发运证据超过730天且缺少经销商/清单更新，自动升级为高置信度账龄异常。",
}


FIELD_ALIASES = {
    "chassis": ["Chassis", "Chassis num", "Chassis_Num", "Chassis Number", "Chassis_Number", "VIN", "VIN Number", "VIN_Number", "Vehicle Identification Number", "Normalized_Chassis"],
    "mismatch_type": ["Mismatch_Type", "Mismatch type", "Type"],
    "found_other": ["Found_In_Other_Lists", "Found in other lists"],
    "so3120": ["SalesOrder_3120", "SO_3120", "Dealer SO", "Dealer_SO"],
    "so3110": ["SalesOrder_3110", "SO_3110", "Factory SO", "Factory_SO"],
    "so_count": ["SO_Count_3120", "SO_Count", "SO Count", "SO_Number_Count"],
    "matnr3120": ["SO_Item0010_MATNR_3120", "MATNR_3120", "Material_3120", "SO_MATNR_3120"],
    "matnr3110": ["SO_Item0010_MATNR_3110", "MATNR_3110", "Material_3110", "SO_MATNR_3110"],
    "pgi": ["SalesOrderPGI_Doc", "SalesOrderPGI_Doc_3120", "PGI_3120", "PGI", "PGI_Doc", "PGI_Doc_3120"],
    "pgi3110": ["SalesOrderPGI_Doc_3110", "PGI_3110", "PGI_Doc_3110", "Factory_PGI"],
    "pgi_date": ["PGI_Date", "PGI date", "PGI_Date_3120", "PGI_Date_3110", "SalesOrderPGI_Date", "SalesOrderPGI_Date_3120", "SalesOrderPGI_Date_3110"],
    "pgi_date3120": ["PGI_Date_3120", "PGI date 3120"],
    "pgi_date3110": ["PGI_Date_3110", "PGI date 3110"],
    "reverse": ["Reverse_PGI", "Reverse_PGI_3120", "Reverse PGI", "Reverse PGI 3120"],
    "reverse3110": ["Reverse_PGI_3110", "Reverse PGI 3110", "Factory Reverse PGI"],
    "last_pgi": ["Last_Movement_Is_PGI", "Last_Movement_Is_PGI_3120", "Last Movement Is PGI"],
    "invoice": ["Invoice_No", "Invoice_No_3120", "Invoice_3120", "Invoice", "Invoice No"],
    "invoice3110": ["Invoice_No_3110", "Invoice_3110", "Factory Invoice"],
    "bill_to": ["BillTo", "Bill-to", "Invoice_BP_Name", "BP Name"],
    "amount": ["BP_Received_Amount", "Received amount", "Amount"],
    "po": ["PO_Number", "PO_Number_3120", "PO", "PO No"],
    "po_count": ["PO_Number_Count", "PO Count", "PO_Count"],
    "gr": ["PO_GR_Date", "GR_Date", "GR Date"],
    "actual": ["Actual situation after check", "Actual situation", "Actual feedback"],
    "dealer_update": ["Dealer_Update_Date", "Dealer update date", "Dealer_Last_Update", "List_Update_Date", "List update date", "Dealer list refresh date"],
}


@dataclass
class DiagnosticResult:
    assumption: str
    judgement: str
    evidence: str
    process_issue: str
    optimisation: str
    owner: str
    priority: str
    confidence: str
    stock_type: str
    lifecycle_stage: str
    issue_category: str = ""
    root_cause: str = ""
    control_gap: str = ""
    recommended_control: str = ""
    required_evidence: str = ""
    next_action: str = ""
    preventive_rule: str = ""
    linked_cases: str = ""
    rule_id: str = ""
    diagnostic_category: str = ""


def cell_text(value):
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    return text


def truthy(value):
    text = cell_text(value).upper()
    return bool(text) and text not in {"N", "NO", "FALSE", "0", "空", "无", "NONE"}


def normalize_chassis(value):
    text = cell_text(value).upper()
    text = re.sub(r"^(VIN|VIN NUMBER|CHASSIS|CHASSIS NUMBER)\s*[:：]\s*", "", text)
    text = re.sub(r"\s*-\s*", "-", text)
    text = re.sub(r"\s+", "", text)
    # Dealer reports sometimes provide a full 17-character VIN while SAP/list matching
    # is performed on chassis number. Use the common last-7 chassis convention so
    # VIN-only inputs can still match chassis-number records.
    if re.fullmatch(r"[A-Z0-9]{17}", text):
        text = text[-7:]
    if re.fullmatch(r"0+\d+", text):
        text = text.lstrip("0") or "0"
    return text


def parse_date(value):
    text = cell_text(value)
    if not text:
        return None
    parsed = pd.to_datetime(text, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime()


def days_since(value, today=None):
    parsed = parse_date(value)
    if parsed is None:
        return None
    today = today or datetime.today()
    return (today.date() - parsed.date()).days


def levenshtein(a, b, max_distance=2):
    if abs(len(a) - len(b)) > max_distance:
        return max_distance + 1
    previous = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        current = [i]
        row_min = i
        for j, cb in enumerate(b, 1):
            cost = 0 if ca == cb else 1
            current.append(min(current[j - 1] + 1, previous[j] + 1, previous[j - 1] + cost))
            row_min = min(row_min, current[-1])
        if row_min > max_distance:
            return max_distance + 1
        previous = current
    return previous[-1]


def first_existing(row, aliases):
    for name in aliases:
        if name in row.index:
            return cell_text(row[name])
    return ""


def find_column(df, aliases):
    normalized = {str(col).strip().lower(): col for col in df.columns}
    for alias in aliases:
        if alias.strip().lower() in normalized:
            return normalized[alias.strip().lower()]
    return None


def count_present(*values):
    return sum(1 for value in values if truthy(value))


def translate_to_chinese(value):
    text = cell_text(value)
    if not text:
        return text
    if text in CHINESE_VALUE_TRANSLATIONS:
        return CHINESE_VALUE_TRANSLATIONS[text]
    if text in CHINESE_PHRASE_TRANSLATIONS:
        return CHINESE_PHRASE_TRANSLATIONS[text]
    translated = text
    # Translate summary prefixes and dynamic evidence labels while preserving IDs, dates and numbers.
    prefix_patterns = [
        ("Rule_ID: ", "规则编号："),
        ("Diagnostic category: ", "诊断类别："),
        ("Data-chain judgement: ", "数据链判断："),
        ("Process issue category: ", "流程问题分类："),
        ("Priority: ", "优先级："),
        ("Action owner: ", "行动负责人："),
        ("Stock_Type: ", "库存类型："),
        ("Confidence: ", "置信度："),
        ("Mismatch: ", "不匹配类型："),
        ("SO count: ", "SO数量："),
        ("Reverse PGI: ", "Reverse PGI："),
        ("Invoice: ", "发票："),
        ("Bill-to: ", "Bill-to："),
        ("Other lists: ", "其他清单："),
        ("PGI age: ", "PGI账龄："),
        ("Dealer update: ", "经销商更新："),
    ]
    for english, chinese in prefix_patterns:
        translated = translated.replace(english, chinese)
    phrase_replacements = {
        " days": "天",
        "Blank": "空白",
        "Only in List": "仅在清单中",
        "Only in SAP": "仅在SAP中",
        "Mismatch detail": "Mismatch明细",
        "Chassis matching": "底盘号匹配",
        "Missing SAP evidence": "SAP证据缺失",
        "Timing / sync delay": "时点/同步延迟",
        "Completed or left stock": "已完成或已离库",
        "PGI without invoice": "PGI无发票",
        "Existence dependency violation": "上下游依赖违规",
        "Invoice master data gap": "发票主数据缺口",
        "PO / GR evidence gap": "PO/GR证据缺口",
        "PGI date sequence exception": "PGI日期顺序异常",
        "Reverse PGI status conflict": "Reverse PGI状态冲突",
        "Material consistency exception": "物料一致性异常",
        "Special stock pool": "特殊库存池",
        "List ownership check": "清单归属检查",
        "Upstream shipped without dealer update": "上游已发运但经销商未更新",
        "Unresolved upstream shipment ageing": "上游发运长期未闭环",
    }
    for english, chinese in {**CHINESE_VALUE_TRANSLATIONS, **phrase_replacements}.items():
        translated = translated.replace(english, chinese)
    return translated


def translate_dataframe_to_chinese(df):
    if df is None or df.empty:
        return df
    translated = df.copy()
    for col in translated.columns:
        if not pd.api.types.is_numeric_dtype(translated[col]):
            translated[col] = translated[col].map(translate_to_chinese)
    translated = translated.rename(columns={col: CHINESE_COLUMN_NAMES.get(col, col) for col in translated.columns})
    return translated


class DiagnosisEngine:
    def __init__(self):
        self.sheets = {}
        self.source_path = None
        self.analysis = pd.DataFrame()
        self.preview = pd.DataFrame()
        self.match_pairs = {}

    def load_workbook(self, path):
        self.source_path = Path(path)
        self.sheets = pd.read_excel(path, sheet_name=None, dtype=object)
        self.sheets = {name.strip(): df for name, df in self.sheets.items()}
        self.preview = self._build_preview()
        return self.preview

    def _build_preview(self):
        rows = []
        for sheet_name, df in self.sheets.items():
            empty_rate = round(float(df.isna().mean().mean()) * 100, 1) if not df.empty else 100.0
            rows.append(
                {
                    "Sheet": sheet_name,
                    "Rows": len(df),
                    "Columns": len(df.columns),
                    "Empty_Rate_%": empty_rate,
                    "Status": "OK" if sheet_name in EXPECTED_SHEETS else "Extra",
                }
            )
        for expected in EXPECTED_SHEETS:
            if expected not in self.sheets:
                rows.append(
                    {
                        "Sheet": expected,
                        "Rows": 0,
                        "Columns": 0,
                        "Empty_Rate_%": 100.0,
                        "Status": "Missing",
                    }
                )
        return pd.DataFrame(rows)

    def run(self):
        detail = self._build_detail_frame()
        if detail.empty:
            raise ValueError("未找到可诊断数据。请确认 Excel 至少包含 Mismatch_Detail、Only_in_List 或 Only_in_SAP。")
        detail = self._attach_normalized_fields(detail)
        self._find_false_positive_pairs(detail)
        results = []
        for _, row in detail.iterrows():
            result = self._diagnose_row(row)
            result = self._complete_process_pack(result)
            output = row.to_dict()
            output.update(
                {
                    "Rule_ID": result.rule_id,
                    "Diagnostic category": result.diagnostic_category,
                    "Normalized_Chassis": normalize_chassis(first_existing(row, FIELD_ALIASES["chassis"])),
                    "Stock_Type": result.stock_type,
                    "Lifecycle_Stage": result.lifecycle_stage,
                    "Actual situation after check": first_existing(row, FIELD_ALIASES["actual"]),
                    "Data-chain judgement": result.judgement,
                    "Evidence-based reason": result.evidence,
                    "Process issues identified": result.process_issue,
                    "How to optimise": result.optimisation,
                    "Process issue category": result.issue_category,
                    "Root cause hypothesis": result.root_cause,
                    "Control gap": result.control_gap,
                    "Recommended control": result.recommended_control,
                    "Required evidence": result.required_evidence,
                    "Next action": result.next_action,
                    "Preventive rule": result.preventive_rule,
                    "Linked example cases": result.linked_cases,
                    "Action owner": result.owner,
                    "Priority": result.priority,
                    "Confidence": result.confidence,
                }
            )
            results.append(output)
        self.analysis = self._order_analysis_columns(pd.DataFrame(results))
        return self.analysis

    def _order_analysis_columns(self, df):
        if df.empty:
            return df
        chassis_column = find_column(df, FIELD_ALIASES["chassis"]) or "Original_Chassis"
        leading = [chassis_column]
        leading.extend([col for col in KEY_COLUMNS_AFTER_CHASSIS if col in df.columns and col not in leading])
        if "Normalized_Chassis" in df.columns and "Normalized_Chassis" not in leading:
            leading.append("Normalized_Chassis")
        for col in ["Data-chain judgement", "Evidence-based reason"]:
            if col in df.columns and col not in leading:
                leading.append(col)
        trailing = [col for col in df.columns if col not in leading]
        return df[leading + trailing]

    def _build_detail_frame(self):
        frames = []
        if "Mismatch_Detail" in self.sheets and not self.sheets["Mismatch_Detail"].empty:
            df = self.sheets["Mismatch_Detail"].copy()
            if find_column(df, FIELD_ALIASES["mismatch_type"]) is None:
                df["Mismatch_Type"] = "Mismatch detail"
            frames.append(df)
        for sheet_name, mismatch_type in [("Only_in_List", "Only in List"), ("Only_in_SAP", "Only in SAP")]:
            if sheet_name in self.sheets and not self.sheets[sheet_name].empty:
                df = self.sheets[sheet_name].copy()
                if find_column(df, FIELD_ALIASES["mismatch_type"]) is None:
                    df["Mismatch_Type"] = mismatch_type
                df["_Source_Sheet"] = sheet_name
                frames.append(df)
        if not frames:
            return pd.DataFrame()
        all_columns = sorted({column for frame in frames for column in frame.columns})
        aligned = [frame.reindex(columns=all_columns) for frame in frames]
        return pd.concat(aligned, ignore_index=True)

    def _complete_process_pack(self, result):
        rule_defaults = {
            "False Positive - Naming": (
                "M2", "Category 1",
                "Chassis naming / matching issue",
                "The two records likely refer to the same physical unit, but manual/SAP formatting created a mismatch.",
                "Mismatch generation happens before robust chassis normalization and fuzzy pairing.",
                "Run normalization and confirmed fuzzy matching before the mismatch list is finalized.",
                "Original chassis, normalized chassis, matched counterpart and matching confidence.",
                "Data team confirms the pair and excludes it from real stock exceptions.",
                "Block true-mismatch classification when normalized chassis matches exactly or approved edit-distance=1 pairing exists.",
            ),
            "SAP No Evidence": (
                "R0", "Category 1",
                "SAP master/lifecycle evidence missing",
                "Dealer list shows the unit, but SAP does not expose enough SO/PGI/Invoice/PO evidence.",
                "Source entry, SAP creation and chassis naming are not reconciled before exception reporting.",
                "Create a missing-SAP-evidence triage path for unbuilt, manually entered or misnamed units.",
                "Original list record, SAP search screenshot/export, PO/SO creation status and normalized chassis check.",
                "SAP support checks whether the unit is uncreated, misnamed or outside SAP-managed stock.",
                "Do not escalate as stock loss until SAP creation and chassis matching checks are completed.",
            ),
            "Process Break": (
                "R3/R1", "Category 1",
                "Lifecycle document break",
                "Downstream movement exists while invoice, reversal or upstream evidence is missing.",
                "The process allows a vehicle to sit between PGI, reversal and invoice without an owner-controlled closure.",
                "Use an exception queue for broken SO/PGI/Invoice/GR chains and require owner sign-off.",
                "SO, PGI, reverse PGI, invoice, GR, handover and movement history.",
                "SAP support reconciles movement chain; Finance checks invoice; Dealer confirms physical status.",
                "Every downstream document must point to its required upstream evidence before the case can close.",
            ),
            "Cancelled Sale / Refund Not Closed": (
                "R11", "Category 1",
                "Return/refund process missing evidence",
                "The unit entered cancellation/refund but money status, invoice state, reversal and stock reinstatement are not visible as a closed package.",
                "No mandatory evidence pack exists for return/refund cases.",
                "Create a return/refund checklist and block closure until all commercial and SAP documents are complete.",
                "Return request, refund approval, payment status, credit note/invoice status, reverse PGI, stock location, dealer list refresh.",
                "Sales admin, Finance and SAP support jointly close the return evidence pack.",
                "Cancellation/refund cases cannot return to normal stock or resale until the release gate is approved.",
            ),
            "PGI Closure Violation": (
                "R12", "Category 1",
                "PGI closure violation",
                "PGI indicates the unit should leave stock, but no departure or reversal proof is visible.",
                "PGI can become the final visible movement without mandatory closure evidence.",
                "Every PGI must be closed by invoice/handover/departure or reverse PGI/return-to-stock proof.",
                "PGI document, invoice/handover status, physical departure confirmation, reverse PGI if returned.",
                "SAP support checks movement; Dealer confirms location; Sales admin confirms handover/reversal.",
                "Create an aging alert for PGI without invoice/handover/reversal beyond the allowed window.",
            ),
            "Reservation Resale Risk": (
                "R13", "Category 1",
                "Reservation resale risk",
                "Multiple reservations or sales attempts appear before the prior customer path is fully closed.",
                "Vehicle release after customer cancellation is not gated by refund, invoice, PGI and stock checks.",
                "Introduce reservation lock and release approval before a unit can be resold.",
                "Reservation history, SO count, customer cancellation, refund status, invoice status, PGI/reversal, release approval.",
                "Sales admin reviews customer paths and confirms the active/closed status before resale.",
                "Block new reservation/SO while cancellation/refund/PGI exceptions remain open.",
            ),
            "Demo / Used Special Handling": (
                "R6", "Category 2",
                "Special stock handling",
                "Used/demo/placeholder stock follows a different lifecycle from normal new-stock sales.",
                "Special stock is mixed into ordinary new-stock reconciliation rules.",
                "Separate Z19, USED and demo units into a dedicated stock pool and reporting path.",
                "MATNR, demo flag, registration/warranty status, physical location and responsible owner.",
                "Sales admin confirms special stock status and removes it from normal-new-stock exception ageing.",
                "Classify special stock before applying normal new-car SO/PGI/Invoice rules.",
            ),
            "Completed / Left Stock": (
                "R9", "Category 1",
                "Completed or left stock",
                "Invoice/PGI evidence suggests the vehicle may already have left the expected stock state.",
                "Dealer list and SAP lifecycle status are not refreshed at the same time.",
                "Require list refresh after invoice/handover/stock departure.",
                "Invoice, handover/departure proof, stock location and dealer list refresh timestamp.",
                "Dealer confirms whether the unit should be removed from the local list.",
                "After invoice or handover, flag stale list records until removed or justified.",
            ),
            "Likely Data Timing Delay": (
                "R8", "Category 2",
                "Data timing / ownership delay",
                "SAP/List evidence points to a timing or ownership update delay rather than confirmed stock loss.",
                "Dealer list refresh, transfer ownership and SAP state updates are not synchronized.",
                "Add refresh timestamp and cross-list ownership checks to the reconciliation report.",
                "Other-list match, transfer record, list refresh timestamp and SAP stock status.",
                "Dealer/Data team confirms latest list refresh and ownership assignment.",
                "Do not close as true mismatch until latest list refresh and ownership checks are complete.",
            ),
            "Needs Manual Review": (
                "C2", "Category 2",
                "Insufficient evidence",
                "The available mapped fields are not enough for an auditable conclusion.",
                "Required lifecycle fields are missing, unmapped or blank.",
                "Collect missing SO, PGI, reverse, invoice, PO/GR and actual feedback fields.",
                "SO, PGI, reverse PGI, invoice, PO/GR, Bill-to, payment status and actual feedback.",
                "Data owner enriches the source workbook and reruns diagnosis.",
                "Records with low evidence confidence should remain Category 2 until required fields are supplied.",
            ),
        }
        default = rule_defaults.get(result.judgement, rule_defaults["Needs Manual Review"])
        (
            rule_id,
            category,
            issue_category,
            root_cause,
            control_gap,
            recommended_control,
            required_evidence,
            next_action,
            preventive_rule,
        ) = default
        result.rule_id = result.rule_id or rule_id
        result.diagnostic_category = result.diagnostic_category or category
        result.issue_category = result.issue_category or issue_category
        result.root_cause = result.root_cause or root_cause
        result.control_gap = result.control_gap or control_gap
        result.recommended_control = result.recommended_control or recommended_control
        result.required_evidence = result.required_evidence or required_evidence
        result.next_action = result.next_action or next_action
        result.preventive_rule = result.preventive_rule or preventive_rule
        return result

    def _attach_normalized_fields(self, df):
        chassis_col = find_column(df, FIELD_ALIASES["chassis"])
        if chassis_col:
            df["Original_Chassis"] = df[chassis_col].map(cell_text)
        else:
            df["Original_Chassis"] = ""
        df["Normalized_Chassis"] = df["Original_Chassis"].map(normalize_chassis)
        return df

    def _find_false_positive_pairs(self, detail):
        only_list = detail[detail.apply(lambda r: "LIST" in first_existing(r, FIELD_ALIASES["mismatch_type"]).upper(), axis=1)]
        sap_like = detail[detail.apply(lambda r: "SAP" in first_existing(r, FIELD_ALIASES["mismatch_type"]).upper(), axis=1)]
        sap_norms = [n for n in sap_like["Normalized_Chassis"].dropna().unique() if n]
        self.match_pairs = {}
        for _, row in only_list.iterrows():
            norm = row["Normalized_Chassis"]
            if not norm:
                continue
            if norm in sap_norms:
                self.match_pairs[norm] = ("exact", norm)
                continue
            for candidate in sap_norms:
                if levenshtein(norm, candidate, 1) <= 1:
                    self.match_pairs[norm] = ("edit_distance_1", candidate)
                    break

    def _diagnose_row(self, row):
        chassis = normalize_chassis(first_existing(row, FIELD_ALIASES["chassis"]))
        mismatch_type = first_existing(row, FIELD_ALIASES["mismatch_type"]) or first_existing(row, ["_Source_Sheet"])
        found_other = first_existing(row, FIELD_ALIASES["found_other"])
        so3120 = first_existing(row, FIELD_ALIASES["so3120"])
        so3110 = first_existing(row, FIELD_ALIASES["so3110"])
        so_count = first_existing(row, FIELD_ALIASES["so_count"])
        matnr = first_existing(row, FIELD_ALIASES["matnr3120"]) or first_existing(row, FIELD_ALIASES["matnr3110"])
        pgi = first_existing(row, FIELD_ALIASES["pgi"])
        pgi_date = first_existing(row, FIELD_ALIASES["pgi_date"])
        dealer_update = first_existing(row, FIELD_ALIASES["dealer_update"])
        reverse = first_existing(row, FIELD_ALIASES["reverse"])
        last_pgi = first_existing(row, FIELD_ALIASES["last_pgi"])
        invoice = first_existing(row, FIELD_ALIASES["invoice"]) or first_existing(row, FIELD_ALIASES["invoice3110"])
        po = first_existing(row, FIELD_ALIASES["po"])
        gr = first_existing(row, FIELD_ALIASES["gr"])
        bill_to = first_existing(row, FIELD_ALIASES["bill_to"])
        amount = first_existing(row, FIELD_ALIASES["amount"])
        actual = first_existing(row, FIELD_ALIASES["actual"])

        material_upper = matnr.upper()
        actual_upper = actual.upper()
        mismatch_upper = mismatch_type.upper()
        evidence_count = count_present(so3120, so3110, pgi, reverse, invoice, po, gr, bill_to, amount)

        if "Z19" in material_upper or "USED" in chassis or "USED" in material_upper:
            stock_type = "Used stock / placeholder"
        elif "DEMO" in actual_upper:
            stock_type = "Demo stock"
        else:
            stock_type = "Normal new stock" if material_upper.startswith("Z12") else "Unclassified stock"

        evidence_bits = []
        for label, value in [
            ("Mismatch", mismatch_type),
            ("SO3120", so3120),
            ("SO3110", so3110),
            ("SO count", so_count),
            ("MATNR", matnr),
            ("PGI", pgi),
            ("PGI date", pgi_date),
            ("Reverse PGI", reverse),
            ("Invoice", invoice),
            ("PO", po),
            ("GR", gr),
            ("Bill-to", bill_to),
            ("Other lists", found_other),
        ]:
            if truthy(value):
                evidence_bits.append(f"{label}: {value}")
        evidence = "; ".join(evidence_bits) if evidence_bits else "No strong lifecycle evidence found."

        confidence = "High" if evidence_count >= 5 else "Medium" if evidence_count >= 2 else "Low"
        try:
            so_count_num = float(so_count) if truthy(so_count) else 0
        except ValueError:
            so_count_num = 0
        has_return_signal = any(word in actual_upper for word in ["REFUND", "RETURN", "CANCEL", "CANCELLED", "CANCELLATION", "退车", "退货", "取消", "退款"])
        has_stock_signal = any(word in actual_upper for word in ["STILL", "STOCK", "NOT SOLD", "DEMO", "库存", "未售", "还在"])
        has_multi_sale_signal = so_count_num >= 2 or any(word in actual_upper for word in ["TWO CUSTOMER", "TWO CUSTOMERS", "MULTIPLE", "SECOND", "RESERV", "二次", "多次", "预定"])

        if chassis in self.match_pairs:
            mode, target = self.match_pairs[chassis]
            assumption = f"Chassis normalization/fuzzy match links this row to SAP/List record {target}."
            return DiagnosticResult(
                assumption,
                "False Positive - Naming",
                evidence,
                "Chassis format or one-character input difference created a false mismatch.",
                "Apply Normalized_Chassis matching before generating mismatch lists.",
                "Data",
                "Low",
                "High" if mode == "exact" else "Medium",
                stock_type,
                "Chassis matching",
            )

        if has_return_signal:
            return DiagnosticResult(
                "Actual feedback indicates sales activity was cancelled or refunded after lifecycle movements occurred.",
                "Cancelled Sale / Refund Not Closed",
                evidence,
                "Return/refund status entered a business exception path, but the data chain does not show a complete evidence package for return approval, refund status, PGI reversal and stock reinstatement.",
                "Create a standard return/refund workflow. Once a unit enters return or cancellation, require return request, refund approval, reverse PGI or stock reinstatement evidence, invoice status confirmation and dealer list refresh before the case can close.",
                "Sales admin",
                "High",
                confidence,
                stock_type,
                "Cancellation / refund loop",
                issue_category="Return/refund process missing evidence",
                root_cause="The transaction moved into refund/cancellation after sales or PGI activity, but the system does not expose enough evidence to tell whether money, invoice, PGI reversal and stock status were closed.",
                control_gap="No mandatory return evidence pack is linked to the stock reconciliation record; PGI becomes the last visible judgement point even when the commercial process may have moved further.",
                recommended_control="Build a return/refund checklist with mandatory fields: return request ID, customer cancellation confirmation, refund approval/status, invoice cancellation or credit note status, reverse PGI or stock reinstatement document, current physical location and list refresh timestamp.",
                required_evidence="Return request, refund approval, payment/refund status, invoice or credit note status, reverse PGI document, stock location confirmation, dealer list refresh evidence.",
                next_action="Sales admin confirms refund stage and collects missing return evidence; Finance confirms money status; SAP support confirms whether reverse PGI or stock reinstatement is required.",
                preventive_rule="If PGI exists and a record is marked cancelled/refunded, block closure until refund, invoice, reverse PGI/stock reinstatement and physical stock evidence are all completed.",
                linked_cases="NG230168; similar cancellation/refund cases",
            )

        if truthy(pgi) and has_stock_signal and not truthy(reverse) and not truthy(invoice):
            return DiagnosticResult(
                "PGI exists while feedback indicates the unit may still remain in stock.",
                "PGI Closure Violation",
                evidence,
                "PGI indicates the vehicle should have left the stock flow, but there is no visible invoice/handover completion or reverse PGI evidence to explain why it remains.",
                "Introduce a PGI closure control: every PGI must end in either confirmed handover/invoice/stock departure or an approved reverse PGI/return-to-stock action.",
                "SAP support",
                "High",
                confidence,
                stock_type,
                "PGI closure exception",
                issue_category="PGI closure violation",
                root_cause="PGI was posted but the downstream evidence does not prove that the vehicle physically left or that the PGI was reversed.",
                control_gap="PGI can remain as the last movement without a mandatory departure or reversal check.",
                recommended_control="Create a PGI exception queue. If PGI is the latest movement and no handover/invoice/reverse PGI appears within the allowed window, assign an owner and block the case from being treated as normal stock.",
                required_evidence="PGI document, handover confirmation, invoice status, physical departure record, or reverse PGI/return-to-stock document.",
                next_action="SAP support checks movement history; Dealer confirms physical stock location; Sales admin confirms whether handover or reversal is missing.",
                preventive_rule="A PGI-posted vehicle must not remain unresolved: it must either leave stock with proof or be reversed back into stock with proof.",
                linked_cases="SRC254700",
            )

        if has_multi_sale_signal:
            return DiagnosticResult(
                "The record suggests repeated reservation/sales attempts or a risk that a reserved unit was released and reused without a clean cancellation gate.",
                "Reservation Resale Risk",
                evidence,
                "Reservation and cancellation controls may not prevent a unit from entering a second sales path before the first process is fully closed.",
                "Add reservation lock and release controls. A reserved vehicle cannot be resold until cancellation, refund, invoice, PGI and stock release checks are complete.",
                "Sales admin",
                "High" if truthy(pgi) else "Medium",
                confidence,
                stock_type,
                "Reservation / resale control",
                issue_category="Reservation resale risk",
                root_cause="Multiple sales attempts or reservation changes appear before the previous customer path is fully closed.",
                control_gap="The system does not enforce a clean release gate between customer cancellation and a new reservation/sale.",
                recommended_control="Use a reservation lock. When a customer cancels or returns, move the unit to an exception status until refund, cancellation approval, invoice status, PGI/reversal and stock availability are confirmed.",
                required_evidence="Reservation history, cancellation confirmation, refund status, SO history, invoice status, PGI/reversal status, stock release approval.",
                next_action="Sales admin reviews all SO/reservation attempts and confirms which customer path is active or closed before the unit is made available again.",
                preventive_rule="Prevent new SO/reservation creation for a unit with open cancellation/refund/PGI exceptions; require release approval before resale.",
                linked_cases="NG230168; SRC256086; SRC263056; SRH256219",
            )

        pgi3110 = first_existing(row, FIELD_ALIASES["pgi3110"])
        reverse3110 = first_existing(row, FIELD_ALIASES["reverse3110"])
        invoice3110 = first_existing(row, FIELD_ALIASES["invoice3110"])
        po_count = first_existing(row, FIELD_ALIASES["po_count"])
        pgi_date3120 = first_existing(row, FIELD_ALIASES["pgi_date3120"])
        pgi_date3110 = first_existing(row, FIELD_ALIASES["pgi_date3110"])
        matnr3110 = first_existing(row, FIELD_ALIASES["matnr3110"])
        matnr3120 = first_existing(row, FIELD_ALIASES["matnr3120"])

        if "Z19199901" in material_upper or (material_upper.startswith("Z19") and not truthy(so3110)):
            return DiagnosticResult(
                "Material indicates a used/placeholder order that has not entered the normal factory flow.",
                "Demo / Used Special Handling",
                evidence,
                "Placeholder or used-stock material is being assessed by normal new-stock sales-chain logic.",
                "Classify Z19/Z19199901 records before normal SO/PGI/Invoice checks and maintain a separate used-stock workflow.",
                "Sales admin",
                "Medium",
                confidence,
                "Used stock / placeholder",
                "Placeholder / initial order",
                rule_id="R6",
                diagnostic_category="Category 2",
            )

        if "ONLY IN LIST" in mismatch_upper and truthy(pgi3110) and truthy(reverse3110):
            return DiagnosticResult(
                "Factory-side PGI was reversed while the dealer list still shows the unit.",
                "Process Break",
                evidence,
                "Factory PGI reversal suggests the vehicle was returned or rolled back, but dealer list/SAP stock state is not aligned.",
                "Add a reversal-close checklist: confirm reverse PGI, physical stock location, dealer list refresh and whether the dealer-side PGI/invoice also needs correction.",
                "SAP support",
                "High",
                confidence,
                stock_type,
                "Factory reversal / dealer list mismatch",
                rule_id="R10",
                diagnostic_category="Category 1",
                issue_category="Reversal return not closed",
                root_cause="Factory movement was reversed, but the dealer-facing stock/list status did not close in the same process.",
                control_gap="Factory reversal and dealer list refresh are not linked as one required close-out flow.",
                recommended_control="Create a reverse-PGI closure checklist covering factory reversal, dealer-side movement, invoice status, physical location and list refresh.",
                required_evidence="Factory PGI, reverse PGI, dealer PGI, invoice status, stock location confirmation and refreshed dealer list.",
                next_action="SAP support validates reverse movement; Dealer confirms physical location; Data team updates list status.",
                preventive_rule="A reverse PGI should automatically create a dealer-list reconciliation task until stock/list ownership is confirmed.",
                linked_cases="NG230168; R10冲销退回 cases",
            )

        if truthy(pgi) and truthy(last_pgi) and not truthy(reverse) and not truthy(invoice):
            return DiagnosticResult(
                "Dealer PGI is the latest visible movement and no dealer invoice is available.",
                "Process Break",
                evidence,
                "PGI has been completed and not reversed, but invoice evidence is missing.",
                "Create an automatic PGI-without-invoice exception. Confirm whether the unit has left stock, invoice is delayed, or PGI should be reversed.",
                "Finance",
                "High",
                confidence,
                stock_type,
                "PGI without invoice",
                rule_id="R3",
                diagnostic_category="Category 1",
                issue_category="PGI after missing invoice",
                root_cause="The data chain shows stock movement but no invoice closure.",
                control_gap="PGI can be posted without invoice follow-up or reversal tracking.",
                recommended_control="Daily exception report for PGI with no invoice beyond allowed SLA; require owner, ageing and closure reason.",
                required_evidence="PGI document, invoice status, handover status, reverse PGI status and current stock location.",
                next_action="Finance checks invoice/credit note; Dealer checks handover; SAP support checks whether movement should be reversed.",
                preventive_rule="PGI without invoice must age into an exception queue until resolved.",
                linked_cases="NG230168; R3 PGI后缺发票 cases",
            )

        if truthy(pgi) and not truthy(so3120):
            return DiagnosticResult(
                "Downstream PGI exists but the expected dealer SO evidence is missing.",
                "Process Break",
                evidence,
                "A downstream document exists without the upstream document normally required by the sales chain.",
                "Audit document lineage and prevent downstream PGI records from being accepted without a linked SO.",
                "SAP support",
                "High",
                confidence,
                stock_type,
                "Existence dependency violation",
                rule_id="R1",
                diagnostic_category="Category 2",
            )

        if truthy(invoice) and not truthy(bill_to):
            return DiagnosticResult(
                "Invoice exists but Bill-to evidence is missing.",
                "Process Break",
                evidence,
                "Billing document exists without a clear billing party.",
                "Make Bill-to mandatory for invoice-linked reconciliation and block closure until the customer account is mapped.",
                "Finance",
                "Medium",
                confidence,
                stock_type,
                "Invoice master data gap",
                rule_id="R7",
                diagnostic_category="Category 2",
            )

        if truthy(so3120) and not truthy(po) and cell_text(po_count) in {"", "0", "0.0"}:
            return DiagnosticResult(
                "Sales order exists but PO/GR evidence is missing or zero.",
                "Likely Data Timing Delay",
                evidence,
                "Procurement evidence is absent even though sales order evidence exists.",
                "Check whether this is internal transfer, delayed PO/GR capture or missing procurement linkage.",
                "Data",
                "Medium",
                confidence,
                stock_type,
                "PO / GR evidence gap",
                rule_id="R8",
                diagnostic_category="Category 2",
            )

        if truthy(pgi_date3120) and truthy(pgi_date3110) and pgi_date3120 < pgi_date3110:
            return DiagnosticResult(
                "Dealer PGI date appears earlier than factory PGI date.",
                "Process Break",
                evidence,
                "The expected factory-to-dealer movement sequence may be reversed or incorrectly dated.",
                "Audit movement timestamps and correct the document sequence if dates were posted out of order.",
                "SAP support",
                "Medium",
                confidence,
                stock_type,
                "PGI date sequence exception",
                rule_id="R2",
                diagnostic_category="Category 2",
            )

        if truthy(reverse) and truthy(last_pgi):
            return DiagnosticResult(
                "Reverse PGI flag and Last_Movement_Is_PGI both appear active.",
                "Process Break",
                evidence,
                "Reversal status conflicts with the last movement status.",
                "Reconcile movement history and ensure reversal status updates the latest movement indicator.",
                "SAP support",
                "High",
                confidence,
                stock_type,
                "Reverse PGI status conflict",
                rule_id="R4",
                diagnostic_category="Category 2",
            )

        upstream_pgi_date = pgi_date3120 or pgi_date3110 or pgi_date
        upstream_pgi_age = days_since(upstream_pgi_date)
        dealer_update_age = days_since(dealer_update)
        stale_dealer_update = not truthy(dealer_update) or (dealer_update_age is not None and dealer_update_age >= 180)
        if (
            "ONLY IN SAP" in mismatch_upper
            and (truthy(pgi) or truthy(pgi3110) or truthy(upstream_pgi_date))
            and upstream_pgi_age is not None
            and upstream_pgi_age >= 730
            and stale_dealer_update
            and not truthy(found_other)
        ):
            aged_evidence = evidence
            age_text = f"PGI age: {upstream_pgi_age} days"
            update_text = f"Dealer update: {dealer_update or 'not available'}"
            aged_evidence = f"{aged_evidence}; {age_text}; {update_text}" if aged_evidence else f"{age_text}; {update_text}"
            return DiagnosticResult(
                "Upstream SAP evidence indicates the vehicle was shipped long ago, but the dealer/list side still has no matching update.",
                "Upstream Shipped / Dealer Not Updated",
                aged_evidence,
                "A vehicle with upstream PGI/shipment evidence has remained absent from dealer-side updates far beyond the normal data refresh window.",
                "Create an ageing control for upstream-shipped vehicles that are still missing from dealer/list updates. Any PGI older than 730 days without dealer confirmation should be escalated as a process-control exception, not treated as ordinary timing delay.",
                "Dealer",
                "High",
                "High",
                stock_type,
                "Upstream shipped without dealer update",
                rule_id="R14",
                diagnostic_category="Category 1",
                issue_category="Unresolved upstream shipment ageing",
                root_cause="The upstream system shows the vehicle was shipped, but dealer-side ownership/list status was never confirmed or refreshed after a long ageing period.",
                control_gap="Long-aged upstream shipment exceptions can remain in a generic timing bucket without a mandatory dealer confirmation SLA.",
                recommended_control="Run a monthly ageing report for Only-in-SAP records with PGI/shipment dates older than 730 days and no dealer list update. Require dealer confirmation, physical location proof and list refresh evidence.",
                required_evidence="PGI/shipment document and date, dealer list refresh history, dealer physical confirmation, ownership/transfer record and final stock status.",
                next_action="Dealer confirms whether the vehicle was received/sold/transferred; Data team checks list refresh history; SAP support validates upstream movement and ownership chain.",
                preventive_rule="If upstream PGI/shipment evidence is older than 730 days and dealer/list update is missing, auto-escalate as a high-confidence ageing exception.",
                linked_cases="R14 upstream-shipped long ageing cases",
            )

        if truthy(matnr3120) and truthy(matnr3110) and matnr3120 != matnr3110 and "Z19" not in matnr3120.upper() and "Z19" not in matnr3110.upper():
            return DiagnosticResult(
                "Dealer and factory material numbers do not match.",
                "Process Break",
                evidence,
                "Material mismatch can point to order creation or vehicle assignment errors.",
                "Validate material mapping between dealer SO and factory SO before closing the reconciliation.",
                "SAP support",
                "Medium",
                confidence,
                stock_type,
                "Material consistency exception",
                rule_id="R5",
                diagnostic_category="Category 2",
            )

        if stock_type in {"Used stock / placeholder", "Demo stock"}:
            return DiagnosticResult(
                "The unit appears to be demo/used/placeholder stock and should not follow normal new-stock logic.",
                "Demo / Used Special Handling",
                evidence,
                "Special stock type is mixed into the ordinary new-stock reconciliation flow.",
                "Separate Z19, USED and demo units into a dedicated review path with explicit demo/used status fields.",
                "Sales admin",
                "Medium",
                confidence,
                stock_type,
                "Special stock pool",
            )

        if truthy(pgi) and truthy(reverse) and not truthy(invoice):
            return DiagnosticResult(
                "PGI and reverse movement exist, but dealer-side invoice evidence is missing.",
                "Process Break",
                evidence,
                "Movement and invoice evidence are inconsistent or incomplete.",
                "Assign owner to reconcile PGI, reversal, invoice and handover status before closing the mismatch.",
                "SAP support",
                "High",
                confidence,
                stock_type,
                "Movement contradiction",
            )

        if "ONLY IN LIST" in mismatch_upper and evidence_count == 0:
            return DiagnosticResult(
                "The vehicle appears on the dealer list, but SAP lifecycle evidence is not available.",
                "SAP No Evidence",
                evidence,
                "SAP master/lifecycle data may be missing, delayed or entered under a different identifier.",
                "Check source entry, SAP creation status and Chassis normalization before treating it as true stock.",
                "SAP support",
                "High",
                "Low",
                stock_type,
                "Missing SAP evidence",
            )

        if "ONLY IN SAP" in mismatch_upper and truthy(found_other):
            return DiagnosticResult(
                "SAP has the vehicle and it is found in another dealer list.",
                "Likely Data Timing Delay",
                evidence,
                "Dealer/list ownership may be stale or updated in one place but not another.",
                "Refresh dealer list ownership and confirm transfer timing before escalating as stock loss.",
                "Dealer",
                "Medium",
                confidence,
                stock_type,
                "List ownership check",
            )

        if truthy(invoice) or (truthy(pgi) and truthy(last_pgi)):
            return DiagnosticResult(
                "Lifecycle evidence suggests the unit may have completed or left the expected stock state.",
                "Completed / Left Stock",
                evidence,
                "List and SAP stock status may be out of sync after invoice, handover or PGI movement.",
                "Confirm handover/stock location and remove or refresh stale dealer list records.",
                "Dealer",
                "Medium",
                confidence,
                stock_type,
                "Completed or left stock",
            )

        if evidence_count >= 2:
            return DiagnosticResult(
                "SAP lifecycle evidence exists, but current List/SAP timing does not fully align.",
                "Likely Data Timing Delay",
                evidence,
                "Data refresh timing or status propagation may be behind the physical stock situation.",
                "Wait for refresh cycle or check the list update mechanism for this dealer.",
                "Data",
                "Low",
                confidence,
                stock_type,
                "Timing / sync delay",
            )

        return DiagnosticResult(
            "The available evidence is not enough to make a confident final judgement.",
            "Needs Manual Review",
            evidence,
            "Key lifecycle fields are missing or not mapped from the source workbook.",
            "Collect SO, PGI, reverse PGI, invoice, PO/GR and actual feedback, then rerun diagnosis.",
            "Data",
            "Medium",
            "Low",
            stock_type,
            "Evidence incomplete",
        )

    def export_report(self, path, language="en"):
        if self.analysis.empty:
            raise ValueError("还没有诊断结果，请先运行诊断。")
        summary = self._summary_frame()
        analysis = self.analysis
        preview = self.preview
        dashboard_sheet = "Dashboard"
        diagnosis_sheet = "Mismatch_Diagnosis"
        quality_sheet = "Import_Quality"
        if language == "zh":
            summary = translate_dataframe_to_chinese(summary)
            analysis = translate_dataframe_to_chinese(analysis)
            preview = translate_dataframe_to_chinese(preview)
            dashboard_sheet = "统计看板"
            diagnosis_sheet = "Mismatch诊断"
            quality_sheet = "导入质量"
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            summary.to_excel(writer, sheet_name=dashboard_sheet, index=False)
            analysis.to_excel(writer, sheet_name=diagnosis_sheet, index=False)
            preview.to_excel(writer, sheet_name=quality_sheet, index=False)
        self._style_report(path, diagnosis_sheet=diagnosis_sheet)

    def _summary_frame(self):
        if self.analysis.empty:
            return pd.DataFrame()
        rows = []
        total = len(self.analysis)
        rows.append({"Metric": "Total mismatch records", "Value": total})
        for column in ["Rule_ID", "Diagnostic category", "Data-chain judgement", "Process issue category", "Priority", "Action owner", "Stock_Type", "Confidence"]:
            if column not in self.analysis.columns:
                continue
            counts = self.analysis[column].fillna("Blank").value_counts()
            for key, value in counts.items():
                rows.append({"Metric": f"{column}: {key}", "Value": int(value)})
        return pd.DataFrame(rows)

    def _style_report(self, path, diagnosis_sheet="Mismatch_Diagnosis"):
        wb = load_workbook(path)
        for ws in wb.worksheets:
            header_fill = PatternFill("solid", fgColor=RULE_COLORS["Header"])
            headers = [c.value for c in ws[1]]
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = Font(bold=True, color="1E2A36")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.freeze_panes = "D2" if ws.title == diagnosis_sheet else "A2"
            ws.auto_filter.ref = ws.dimensions
            for row in ws.iter_rows(min_row=2):
                priority = ""
                if ws.title == diagnosis_sheet:
                    if "Priority" in headers:
                        priority = row[headers.index("Priority")].value
                    elif "优先级" in headers:
                        priority = {"高": "High", "中": "Medium", "低": "Low"}.get(row[headers.index("优先级")].value, "")
                fill_color = RULE_COLORS.get(priority, None)
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    if fill_color:
                        cell.fill = PatternFill("solid", fgColor=fill_color)
            column_widths = {}
            for col_idx, column_cells in enumerate(ws.columns, 1):
                header = ws.cell(1, col_idx).value
                values = [cell_text(cell.value) for cell in column_cells[:60]]
                width = min(max(12, max((len(v) for v in values), default=12) + 2), 42)
                if header in {
                    "Actual situation after check",
                    "Evidence-based reason",
                    "Process issues identified",
                    "How to optimise",
                    "Root cause hypothesis",
                    "Control gap",
                    "Recommended control",
                    "Required evidence",
                    "Next action",
                    "Preventive rule",
                }:
                    width = 62 if header in {"Process issues identified", "How to optimise"} else 54
                elif header in {"Chassis", "Chassis num", "Original_Chassis", "Normalized_Chassis"}:
                    width = 22
                elif header in {"Priority", "Confidence"}:
                    width = 13
                elif header in {"Rule_ID", "Diagnostic category"}:
                    width = 16
                elif header in {"Action owner", "Stock_Type", "Lifecycle_Stage", "Mismatch_Type", "Data-chain judgement", "Process issue category", "Linked example cases"}:
                    width = 24
                column_widths[col_idx] = width
                ws.column_dimensions[get_column_letter(col_idx)].width = width
            ws.row_dimensions[1].height = 30
            for row in ws.iter_rows(min_row=2):
                estimated_lines = 1
                for cell in row:
                    text = cell_text(cell.value)
                    if not text:
                        continue
                    width = max(column_widths.get(cell.column, 18), 8)
                    chars_per_line = max(int(width * 1.15), 8)
                    line_count = sum(max(1, math.ceil(len(part) / chars_per_line)) for part in text.splitlines() or [""])
                    estimated_lines = max(estimated_lines, line_count)
                ws.row_dimensions[row[0].row].height = min(max(24, 16 + estimated_lines * 15), 132)
        wb.save(path)


class SpreadsheetGrid:
    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        self.frame = ttk.Frame(parent)
        self.frame.grid(row=0, column=0, sticky="nsew")
        self.frame.rowconfigure(1, weight=1)
        self.frame.columnconfigure(0, weight=1)

        self.toolbar = ttk.Frame(self.frame)
        self.toolbar.grid(row=0, column=0, columnspan=2, sticky="ew", padx=8, pady=(8, 4))
        ttk.Button(self.toolbar, text="<", width=4, style="Ghost.TButton", command=lambda: self.canvas.xview_scroll(-8, "units")).pack(side=LEFT, padx=(0, 6))
        ttk.Button(self.toolbar, text=">", width=4, style="Ghost.TButton", command=lambda: self.canvas.xview_scroll(8, "units")).pack(side=LEFT, padx=(0, 10))
        ttk.Button(self.toolbar, text="复制选区", style="Ghost.TButton", command=self.copy_selection).pack(side=LEFT, padx=(0, 10))
        ttk.Button(self.toolbar, text="清除筛选", style="Ghost.TButton", command=self.clear_filters).pack(side=LEFT)
        ttk.Label(self.toolbar, text="拖拽蓝色框选择任意单元格；Ctrl+C 复制；双击筛选行设置列过滤").pack(side=LEFT, padx=12)

        self.canvas = tk.Canvas(self.frame, bg="#FFFFFF", highlightthickness=0)
        self.v_scroll = ttk.Scrollbar(self.frame, orient="vertical", command=self.canvas.yview)
        self.h_scroll = ttk.Scrollbar(self.frame, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=self.v_scroll.set, xscrollcommand=self.h_scroll.set)
        self.canvas.grid(row=1, column=0, sticky="nsew")
        self.v_scroll.grid(row=1, column=1, sticky="ns")
        self.h_scroll.grid(row=2, column=0, sticky="ew")

        self.df = pd.DataFrame()
        self.display = pd.DataFrame()
        self.columns = []
        self.col_widths = []
        self.row_heights = []
        self.filters = {}
        self.header_h = 34
        self.filter_h = 30
        self.default_row_h = 40
        self.selection_anchor = None
        self.selection_current = None

        self.canvas.bind("<Button-1>", self.on_click)
        self.canvas.bind("<B1-Motion>", self.on_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_release)
        self.canvas.bind("<Double-Button-1>", self.on_double_click)
        self.canvas.bind("<Control-c>", lambda _event: self.copy_selection())
        self.canvas.bind("<Control-C>", lambda _event: self.copy_selection())

    def set_dataframe(self, df, hidden_columns=None):
        if hidden_columns:
            df = df[[col for col in df.columns if col not in hidden_columns]]
        self.df = df.copy()
        self.filters = {col: self.filters.get(col, "") for col in self.df.columns}
        self._apply_filters()

    def _apply_filters(self):
        display = self.df.copy()
        for col, value in self.filters.items():
            value = value.strip().lower()
            if value and col in display.columns:
                display = display[display[col].map(lambda item: value in cell_text(item).lower())]
        self.display = display.head(700)
        self.columns = [str(col) for col in self.display.columns]
        self.col_widths = [self._smart_width(col) for col in self.display.columns]
        self.row_heights = [self._smart_row_height(row) for _, row in self.display.iterrows()]
        self.selection_anchor = None
        self.selection_current = None
        self.draw()

    def _smart_width(self, col):
        lower = str(col).lower()
        if lower in {"process issues identified", "how to optimise", "root cause hypothesis", "control gap", "recommended control", "required evidence", "next action", "preventive rule"}:
            return 360
        if lower in {"chassis", "chassis num", "normalized_chassis", "original_chassis"}:
            return 170
        if lower in {"priority", "confidence"}:
            return 90
        if lower in {"action owner", "stock_type", "lifecycle_stage", "mismatch_type", "process issue category"}:
            return 170
        samples = [str(col)] + [cell_text(v) for v in self.display[col].head(40)] if col in self.display.columns else [str(col)]
        max_len = max((len(v) for v in samples), default=10)
        return max(80, min(240, max_len * 8 + 26))

    def _smart_row_height(self, row):
        lines = 1
        for col, width in zip(self.display.columns, self.col_widths):
            text = cell_text(row[col])
            if not text:
                continue
            if str(col).lower() in {"process issues identified", "how to optimise", "root cause hypothesis", "control gap", "recommended control", "required evidence", "next action", "preventive rule"}:
                chars = max(int(width / 7), 18)
                lines = max(lines, min(5, math.ceil(len(text) / chars)))
        return max(self.default_row_h, min(112, 18 + lines * 19))

    def draw(self):
        self.canvas.delete("all")
        x = 0
        total_h = self.header_h + self.filter_h + sum(self.row_heights)
        for idx, col in enumerate(self.columns):
            width = self.col_widths[idx]
            self.canvas.create_rectangle(x, 0, x + width, self.header_h, fill="#EDF2F5", outline="#D6DEE5")
            self.canvas.create_text(x + 8, self.header_h / 2, text=col, anchor="w", font=("Microsoft YaHei UI", 9, "bold"), fill="#1E2A36", width=width - 16)
            filter_text = self.filters.get(self.display.columns[idx], "")
            self.canvas.create_rectangle(x, self.header_h, x + width, self.header_h + self.filter_h, fill="#F8FAFB", outline="#E0E6EA")
            self.canvas.create_text(x + 8, self.header_h + self.filter_h / 2, text=filter_text or "Filter...", anchor="w", font=("Microsoft YaHei UI", 8), fill="#5C6B75" if filter_text else "#A0AAB2", width=width - 16)
            x += width

        y = self.header_h + self.filter_h
        selected = self._selection_bounds()
        for r, (_, row) in enumerate(self.display.iterrows()):
            row_h = self.row_heights[r]
            x = 0
            fill = "#FFFFFF" if r % 2 == 0 else "#FBFCFD"
            for c, col in enumerate(self.display.columns):
                width = self.col_widths[c]
                is_selected = selected and selected[0] <= r <= selected[2] and selected[1] <= c <= selected[3]
                self.canvas.create_rectangle(x, y, x + width, y + row_h, fill="#D7E8EF" if is_selected else fill, outline="#E9EEF2")
                self.canvas.create_text(x + 8, y + 9, text=cell_text(row[col]), anchor="nw", font=("Microsoft YaHei UI", 9), fill="#1F2D35", width=width - 16)
                x += width
            y += row_h
        self.canvas.configure(scrollregion=(0, 0, sum(self.col_widths), total_h))

    def on_click(self, event):
        self.canvas.focus_set()
        cell = self._event_to_cell(event)
        if cell:
            self.selection_anchor = cell
            self.selection_current = cell
            self.draw()

    def on_drag(self, event):
        cell = self._event_to_cell(event)
        if cell and self.selection_anchor:
            self.selection_current = cell
            self.draw()

    def on_release(self, _event):
        return "break"

    def on_double_click(self, event):
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        if not (self.header_h <= y <= self.header_h + self.filter_h):
            return
        col_idx = self._x_to_col(x)
        if col_idx is None:
            return
        col = self.df.columns[col_idx]
        current = self.filters.get(col, "")
        value = simpledialog.askstring("Column filter", f"Filter {col} contains:", initialvalue=current, parent=self.app)
        if value is not None:
            self.filters[col] = value.strip()
            self._apply_filters()

    def clear_filters(self):
        self.filters = {col: "" for col in self.df.columns}
        self._apply_filters()
        self.app.status.set("已清除诊断结果筛选")

    def copy_selection(self):
        bounds = self._selection_bounds()
        if not bounds:
            self.app.status.set("请先拖拽选择要复制的单元格区域")
            return "break"
        r1, c1, r2, c2 = bounds
        lines = []
        for r in range(r1, r2 + 1):
            values = []
            for c in range(c1, c2 + 1):
                values.append(cell_text(self.display.iloc[r, c]))
            lines.append("\t".join(values))
        self.app.clipboard_clear()
        self.app.clipboard_append("\n".join(lines))
        self.app.status.set(f"已复制 {r2 - r1 + 1} 行 x {c2 - c1 + 1} 列")
        return "break"

    def _event_to_cell(self, event):
        x = self.canvas.canvasx(event.x)
        y = self.canvas.canvasy(event.y)
        if y < self.header_h + self.filter_h:
            return None
        col = self._x_to_col(x)
        if col is None:
            return None
        row = self._y_to_row(y)
        if row is None:
            return None
        return (row, col)

    def _x_to_col(self, x):
        pos = 0
        for idx, width in enumerate(self.col_widths):
            if pos <= x < pos + width:
                return idx
            pos += width
        return None

    def _y_to_row(self, y):
        pos = self.header_h + self.filter_h
        for idx, height in enumerate(self.row_heights):
            if pos <= y < pos + height:
                return idx
            pos += height
        return None

    def _selection_bounds(self):
        if not self.selection_anchor or not self.selection_current:
            return None
        r1, c1 = self.selection_anchor
        r2, c2 = self.selection_current
        return (min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2))


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1460x860")
        self.minsize(1240, 720)
        self.engine = DiagnosisEngine()
        self.current_file = tk.StringVar(value="尚未导入文件")
        self.status = tk.StringVar(value="准备就绪")
        self.kpis = {}
        self.rules_df = pd.DataFrame()
        self._configure_theme()
        self._build_layout()

    def _configure_theme(self):
        self.configure(bg="#F4F6F8")
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TFrame", background="#F4F6F8")
        style.configure("Sidebar.TFrame", background="#10202B")
        style.configure("TLabel", background="#F4F6F8", foreground="#24313A", font=("Microsoft YaHei UI", 10))
        style.configure("Title.TLabel", background="#10202B", foreground="#FFFFFF", font=("Microsoft YaHei UI", 18, "bold"))
        style.configure("Subtle.TLabel", background="#10202B", foreground="#AEBCC6", font=("Microsoft YaHei UI", 9))
        style.configure("Card.TFrame", background="#FFFFFF", relief="flat", borderwidth=0)
        style.configure("CardTitle.TLabel", background="#FFFFFF", foreground="#60717D", font=("Microsoft YaHei UI", 9))
        style.configure("CardValue.TLabel", background="#FFFFFF", foreground="#14212B", font=("Microsoft YaHei UI", 18, "bold"))
        style.configure("Accent.TButton", font=("Microsoft YaHei UI", 10, "bold"), padding=(14, 9), background="#246B7A", foreground="#FFFFFF", borderwidth=0, focusthickness=0)
        style.map("Accent.TButton", background=[("active", "#1D5966"), ("pressed", "#184B55")])
        style.configure("Secondary.TButton", font=("Microsoft YaHei UI", 10), padding=(12, 8), background="#E6EBEF", foreground="#1F2D35", borderwidth=0, focusthickness=0)
        style.map("Secondary.TButton", background=[("active", "#D8E0E6"), ("pressed", "#CCD6DE")])
        style.configure("Ghost.TButton", font=("Microsoft YaHei UI", 9), padding=(8, 5), background="#F4F6F8", foreground="#30424E", borderwidth=0, focusthickness=0)
        style.map("Ghost.TButton", background=[("active", "#E7EDF1"), ("pressed", "#DCE5EB")])
        style.configure("TButton", font=("Microsoft YaHei UI", 10), padding=(10, 8), background="#E6EBEF", foreground="#1F2D35", borderwidth=0)
        style.configure("Treeview", rowheight=40, font=("Microsoft YaHei UI", 9), background="#FFFFFF", fieldbackground="#FFFFFF", foreground="#1F2D35", borderwidth=0)
        style.map("Treeview", background=[("selected", "#D7E8EF")], foreground=[("selected", "#10202B")])
        style.configure("Treeview.Heading", font=("Microsoft YaHei UI", 9, "bold"), background="#EDF2F5", foreground="#1E2A36", borderwidth=0, padding=(8, 8))
        style.configure("TNotebook", background="#F4F6F8", borderwidth=0)
        style.configure("TNotebook.Tab", font=("Microsoft YaHei UI", 10), padding=(18, 9), background="#E2E6EA", borderwidth=0)
        style.map("TNotebook.Tab", background=[("selected", "#FFFFFF")], foreground=[("selected", "#12212B")])

    def _build_layout(self):
        sidebar = ttk.Frame(self, style="Sidebar.TFrame", width=300)
        sidebar.pack(side=LEFT, fill=Y)
        sidebar.pack_propagate(False)

        ttk.Label(sidebar, text="Mismatch 诊断", style="Title.TLabel").pack(anchor="w", padx=24, pady=(26, 4))
        ttk.Label(sidebar, text="数据链 + 实际反馈 + 业务状态判断", style="Subtle.TLabel").pack(anchor="w", padx=24)
        ttk.Separator(sidebar).pack(fill=X, padx=24, pady=22)

        ttk.Button(sidebar, text="导入 Excel", style="Accent.TButton", command=self.import_file).pack(fill=X, padx=24, pady=(0, 10))
        ttk.Button(sidebar, text="运行智能诊断", style="Secondary.TButton", command=self.run_diagnosis).pack(fill=X, padx=24, pady=10)
        ttk.Button(sidebar, text="导出分析报告", style="Secondary.TButton", command=self.export_report).pack(fill=X, padx=24, pady=10)
        ttk.Button(sidebar, text="导出中文报告", style="Secondary.TButton", command=self.export_chinese_report).pack(fill=X, padx=24, pady=10)
        ttk.Button(sidebar, text="写入反馈与改进", style="Secondary.TButton", command=self.save_feedback).pack(fill=X, padx=24, pady=10)

        ttk.Separator(sidebar).pack(fill=X, padx=24, pady=22)
        ttk.Label(sidebar, text="当前文件", style="Subtle.TLabel").pack(anchor="w", padx=24)
        ttk.Label(sidebar, textvariable=self.current_file, style="Subtle.TLabel", wraplength=230).pack(anchor="w", padx=24, pady=(6, 22))
        ttk.Label(sidebar, text="状态", style="Subtle.TLabel").pack(anchor="w", padx=24)
        ttk.Label(sidebar, textvariable=self.status, style="Subtle.TLabel", wraplength=230).pack(anchor="w", padx=24, pady=(6, 0))

        main = ttk.Frame(self)
        main.pack(side=RIGHT, expand=True, fill=BOTH)

        header = ttk.Frame(main)
        header.pack(side=TOP, fill=X, padx=22, pady=(20, 12))
        ttk.Label(header, text=APP_TITLE, font=("Microsoft YaHei UI", 22, "bold")).pack(anchor="w")
        ttk.Label(header, text="导入后自动诊断，优先展示流程问题和优化动作；报告导出为可转发 Excel。").pack(anchor="w", pady=(4, 0))

        cards = ttk.Frame(main)
        cards.pack(fill=X, padx=22, pady=(0, 12))
        for label in ["总记录", "High", "Medium", "Low", "需人工确认"]:
            card = ttk.Frame(cards, style="Card.TFrame")
            card.pack(side=LEFT, expand=True, fill=X, padx=(0, 10))
            ttk.Label(card, text=label, style="CardTitle.TLabel").pack(anchor="w", padx=16, pady=(12, 2))
            value = ttk.Label(card, text="0", style="CardValue.TLabel")
            value.pack(anchor="w", padx=16, pady=(0, 12))
            self.kpis[label] = value

        self.notebook = ttk.Notebook(main)
        self.notebook.pack(expand=True, fill=BOTH, padx=22, pady=(0, 18))

        self.preview_tree = self._create_tab("数据预览")
        self.analysis_tree = self._create_tab("诊断结果", spreadsheet=True)
        self.dashboard_tree = self._create_tab("统计看板")
        self.rules_tree = self._create_tab("规则说明")
        self.feedback_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.feedback_frame, text="反馈录入")
        self._build_feedback_tab()
        self._fill_rules()
        self._build_rules_editor(self.rules_tree.master)

    def _create_tab(self, title, spreadsheet=False):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text=title)
        if spreadsheet:
            frame.rowconfigure(0, weight=1)
            frame.columnconfigure(0, weight=1)
            return SpreadsheetGrid(frame, self)
        frame.rowconfigure(1, weight=1)
        frame.columnconfigure(0, weight=1)

        toolbar = ttk.Frame(frame)
        toolbar.grid(row=0, column=0, columnspan=2, sticky="ew", padx=8, pady=(8, 4))
        tree = ttk.Treeview(frame, show="headings", selectmode="extended")
        y_scroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        x_scroll = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        ttk.Button(toolbar, text="<", width=4, style="Ghost.TButton", command=lambda: tree.xview_scroll(-8, "units")).pack(side=LEFT, padx=(0, 6))
        ttk.Button(toolbar, text=">", width=4, style="Ghost.TButton", command=lambda: tree.xview_scroll(8, "units")).pack(side=LEFT, padx=(0, 10))
        ttk.Button(toolbar, text="复制选中", style="Ghost.TButton", command=lambda: self.copy_tree_selection(tree)).pack(side=LEFT)
        ttk.Label(toolbar, text="可多选行并 Ctrl+C 复制；底部滚动条支持横向查看").pack(side=LEFT, padx=12)
        tree.bind("<Control-c>", lambda event: self.copy_tree_selection(tree))
        tree.bind("<Control-C>", lambda event: self.copy_tree_selection(tree))

        tree.grid(row=1, column=0, sticky="nsew")
        y_scroll.grid(row=1, column=1, sticky="ns")
        x_scroll.grid(row=2, column=0, sticky="ew")
        return tree

    def _build_feedback_tab(self):
        panel = ttk.Frame(self.feedback_frame, style="Card.TFrame")
        panel.pack(fill=BOTH, expand=True, padx=18, pady=18)

        top = ttk.Frame(panel, style="Card.TFrame")
        top.pack(fill=X, padx=18, pady=(18, 12))
        ttk.Label(top, text="选择 Chassis", background="#FFFFFF").pack(side=LEFT)
        self.feedback_chassis = ttk.Combobox(top, width=36, state="readonly")
        self.feedback_chassis.pack(side=LEFT, padx=(10, 16))
        ttk.Button(top, text="载入", style="Secondary.TButton", command=self.load_feedback_row).pack(side=LEFT)
        ttk.Button(top, text="写入结果", style="Accent.TButton", command=self.save_feedback).pack(side=LEFT, padx=10)

        ttk.Label(panel, text="Actual situation after check", background="#FFFFFF", font=("Microsoft YaHei UI", 10, "bold")).pack(anchor="w", padx=18)
        self.actual_text = tk.Text(panel, height=8, font=("Microsoft YaHei UI", 10), wrap="word", relief="flat", bg="#F8FAFB", padx=10, pady=10)
        self.actual_text.pack(fill=BOTH, expand=True, padx=18, pady=(6, 14))

        ttk.Label(panel, text="手动输入流程化改进建议 / How to optimise", background="#FFFFFF", font=("Microsoft YaHei UI", 10, "bold")).pack(anchor="w", padx=18)
        self.optimise_text = tk.Text(panel, height=8, font=("Microsoft YaHei UI", 10), wrap="word", relief="flat", bg="#F8FAFB", padx=10, pady=10)
        self.optimise_text.pack(fill=BOTH, expand=True, padx=18, pady=(6, 18))

    def _fill_rules(self):
        rows = [
            ("M2", "False Positive - Naming", "Low", "Data", "标准化一致或编辑距离=1，优先排除假 mismatch", "Apply Normalized_Chassis matching before generating mismatch lists."),
            ("R0", "SAP No Evidence", "High", "SAP support", "Only in List 且 SO/PGI/Invoice/PO-GR 均无证据", "Check SAP creation status, source entry and chassis normalization."),
            ("R1", "Existence dependency violation", "High", "SAP support", "有下游单据但缺上游单据，如有 PGI 但无 SO", "Audit document lineage and prevent downstream documents without upstream evidence."),
            ("R2", "PGI date sequence exception", "Medium", "SAP support", "PGI_Date_3120 早于 PGI_Date_3110", "Audit movement timestamps and correct out-of-order posting."),
            ("R3", "PGI after missing invoice", "High", "Finance", "Last_Movement_Is_PGI=Y 且未冲销，但 Invoice 为空", "Create PGI-without-invoice exception queue with owner and ageing."),
            ("R4", "Reverse PGI status conflict", "High", "SAP support", "Reverse_PGI=Y 但 Last_Movement_Is_PGI=Y", "Reconcile movement history and latest movement indicator."),
            ("R5", "Material consistency exception", "Medium", "SAP support", "MATNR_3120 与 MATNR_3110 不一致，排除 Z19", "Validate material mapping between dealer SO and factory SO."),
            ("R6", "Demo / Used Special Handling", "Medium", "Sales admin", "Z19、Z19199901、USED、demo 单独处理", "Separate special stock from ordinary new-stock reconciliation."),
            ("R7", "Invoice Bill-to gap", "Medium", "Finance", "有 Invoice 但无 Bill-to", "Make Bill-to mandatory for invoice-linked reconciliation."),
            ("R8", "PO / GR evidence gap", "Medium", "Data", "有 SO 但 PO/GR 缺失或 PO count 为 0", "Check internal transfer, delayed PO/GR capture or missing procurement linkage."),
            ("R9", "Completed / Left Stock", "Medium", "Dealer", "Only in SAP 且 invoice/handover/PGI 证据显示可能已完成或离场", "Confirm handover/location and refresh stale dealer list records."),
            ("R10", "Reversal return not closed", "High", "SAP support", "Only in List 且 3110 PGI 已冲销", "Close factory reversal, dealer movement, invoice and dealer list status together."),
            ("R11", "Return/refund process missing evidence", "High", "Sales admin", "退车/退款/取消后缺 return、refund、reverse PGI、stock reinstatement 凭证", "Build a mandatory return/refund evidence pack."),
            ("R12", "PGI Closure Violation", "High", "SAP support", "PGI 后车辆仍在库存或缺离场/回滚证明", "Every PGI must close by departure proof or reverse PGI proof."),
            ("R13", "Reservation Resale Risk", "High", "Sales admin", "多次预定/销售尝试，取消释放门控不清晰", "Use reservation lock and release approval before resale."),
            ("R14", "Upstream Shipped / Dealer Not Updated", "High", "Dealer", "Only in SAP，PGI/发运日期超过730天且经销商清单长期无更新", "Escalate long-aged upstream shipment records as high-confidence ageing exceptions."),
            ("C2", "Needs Manual Review", "Medium", "Data", "字段不足或证据低置信，需要人工补充", "Collect required lifecycle evidence and rerun diagnosis."),
        ]
        self.rules_df = pd.DataFrame(rows, columns=["Rule_ID", "Rule", "Priority", "Owner", "Meaning", "Suggested optimisation"])
        self._show_dataframe(self.rules_tree, self.rules_df)
        self.rules_tree.bind("<<TreeviewSelect>>", self.load_selected_rule)

    def _build_rules_editor(self, frame):
        editor = ttk.Frame(frame, style="Card.TFrame")
        editor.grid(row=3, column=0, columnspan=2, sticky="ew", padx=8, pady=(8, 10))
        for col in range(8):
            editor.columnconfigure(col, weight=1)

        self.rule_id_var = tk.StringVar()
        self.rule_name_var = tk.StringVar()
        self.rule_priority_var = tk.StringVar(value="Medium")
        self.rule_owner_var = tk.StringVar()
        self.rule_meaning_var = tk.StringVar()
        self.rule_optimise_var = tk.StringVar()

        fields = [
            ("Rule ID", self.rule_id_var, 0, 0, 10),
            ("Rule", self.rule_name_var, 0, 2, 24),
            ("Priority", self.rule_priority_var, 0, 4, 12),
            ("Owner", self.rule_owner_var, 0, 6, 16),
            ("Meaning", self.rule_meaning_var, 1, 0, 42),
            ("Suggested optimisation", self.rule_optimise_var, 1, 4, 42),
        ]
        for label, var, row, col, width in fields:
            ttk.Label(editor, text=label, background="#FFFFFF").grid(row=row * 2, column=col, columnspan=2, sticky="w", padx=10, pady=(10, 2))
            if label == "Priority":
                widget = ttk.Combobox(editor, textvariable=var, values=["High", "Medium", "Low"], width=width, state="readonly")
            else:
                widget = ttk.Entry(editor, textvariable=var, width=width)
            widget.grid(row=row * 2 + 1, column=col, columnspan=2, sticky="ew", padx=10, pady=(0, 8))

        buttons = ttk.Frame(editor, style="Card.TFrame")
        buttons.grid(row=4, column=0, columnspan=8, sticky="e", padx=10, pady=(0, 10))
        ttk.Button(buttons, text="新增规则", style="Secondary.TButton", command=self.add_rule).pack(side=LEFT, padx=6)
        ttk.Button(buttons, text="更新选中规则", style="Accent.TButton", command=self.update_rule).pack(side=LEFT, padx=6)

    def load_selected_rule(self, _event=None):
        selected = self.rules_tree.selection()
        if not selected:
            return
        values = self.rules_tree.item(selected[0], "values")
        if len(values) < 6:
            return
        self.rule_id_var.set(values[0])
        self.rule_name_var.set(values[1])
        self.rule_priority_var.set(values[2])
        self.rule_owner_var.set(values[3])
        self.rule_meaning_var.set(values[4])
        self.rule_optimise_var.set(values[5])

    def add_rule(self):
        row = self._rule_form_row()
        if not row["Rule"]:
            messagebox.showwarning("规则不完整", "请至少填写 Rule。")
            return
        if not row["Rule_ID"]:
            row["Rule_ID"] = str(len(self.rules_df) + 1)
        self.rules_df = pd.concat([self.rules_df, pd.DataFrame([row])], ignore_index=True)
        self._show_dataframe(self.rules_tree, self.rules_df)
        self.status.set("已新增规则说明")

    def update_rule(self):
        selected = self.rules_tree.selection()
        if not selected:
            messagebox.showwarning("未选择规则", "请先在规则表中选择一条规则。")
            return
        item_index = self.rules_tree.index(selected[0])
        row = self._rule_form_row()
        for col, value in row.items():
            self.rules_df.loc[item_index, col] = value
        self._show_dataframe(self.rules_tree, self.rules_df)
        self.status.set("已更新选中规则")

    def _rule_form_row(self):
        return {
            "Rule_ID": self.rule_id_var.get().strip(),
            "Rule": self.rule_name_var.get().strip(),
            "Priority": self.rule_priority_var.get().strip() or "Medium",
            "Owner": self.rule_owner_var.get().strip(),
            "Meaning": self.rule_meaning_var.get().strip(),
            "Suggested optimisation": self.rule_optimise_var.get().strip(),
        }

    def import_file(self):
        path = filedialog.askopenfilename(
            title="选择对账 Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            preview = self.engine.load_workbook(path)
            self.current_file.set(Path(path).name)
            self.status.set("导入完成，正在自动运行智能诊断")
            self._show_dataframe(self.preview_tree, preview)
            self._show_dataframe(self.dashboard_tree, preview)
            self.run_diagnosis()
        except Exception as exc:
            messagebox.showerror("导入失败", str(exc))
            self.status.set("导入失败")

    def run_diagnosis(self):
        try:
            analysis = self.engine.run()
            self._show_dataframe(self.analysis_tree, analysis, hidden_columns=DISPLAY_HIDDEN_COLUMNS)
            self._show_dataframe(self.dashboard_tree, self.engine._summary_frame())
            self._update_kpis()
            chassis_values = analysis["Normalized_Chassis"].dropna().astype(str).tolist()
            self.feedback_chassis["values"] = chassis_values
            self.status.set("诊断完成，可以录入反馈或导出报告")
            self.notebook.select(1)
        except Exception as exc:
            messagebox.showerror("诊断失败", str(exc))
            self.status.set("诊断失败")

    def export_report(self):
        self._export_report(language="en")

    def export_chinese_report(self):
        self._export_report(language="zh")

    def _export_report(self, language="en"):
        if self.engine.analysis.empty:
            messagebox.showwarning("尚无结果", "请先运行智能诊断。")
            return
        suffix = "中文诊断报告" if language == "zh" else "智能诊断报告"
        title = "导出中文分析报告" if language == "zh" else "导出分析报告"
        default_name = f"{self.engine.source_path.stem}_{suffix}.xlsx" if self.engine.source_path else f"{suffix}.xlsx"
        path = filedialog.asksaveasfilename(
            title=title,
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel workbook", "*.xlsx")],
        )
        if not path:
            return
        try:
            self.engine.export_report(path, language=language)
            self.status.set(f"报告已导出：{Path(path).name}")
            messagebox.showinfo("导出完成", f"报告已生成：\n{path}")
        except Exception as exc:
            messagebox.showerror("导出失败", str(exc))

    def load_feedback_row(self):
        chassis = self.feedback_chassis.get()
        if not chassis or self.engine.analysis.empty:
            return
        match = self.engine.analysis[self.engine.analysis["Normalized_Chassis"].astype(str) == chassis]
        self.actual_text.delete("1.0", END)
        self.optimise_text.delete("1.0", END)
        if not match.empty:
            self.actual_text.insert("1.0", cell_text(match.iloc[0].get("Actual situation after check", "")))
            self.optimise_text.insert("1.0", cell_text(match.iloc[0].get("How to optimise", "")))

    def save_feedback(self):
        chassis = self.feedback_chassis.get()
        if not chassis or self.engine.analysis.empty:
            messagebox.showwarning("无法保存", "请先运行诊断，并在反馈页选择 Chassis。")
            return
        actual = self.actual_text.get("1.0", END).strip()
        optimisation = self.optimise_text.get("1.0", END).strip()
        mask = self.engine.analysis["Normalized_Chassis"].astype(str) == chassis
        self.engine.analysis.loc[mask, "Actual situation after check"] = actual
        self.engine.analysis.loc[mask, "How to optimise"] = optimisation
        self._show_dataframe(self.analysis_tree, self.engine.analysis, hidden_columns=DISPLAY_HIDDEN_COLUMNS)
        self.status.set("实际情况和流程改进已写入当前诊断结果，导出报告时会包含该内容")

    def _update_kpis(self):
        analysis = self.engine.analysis
        self.kpis["总记录"].configure(text=str(len(analysis)))
        for label in ["High", "Medium", "Low"]:
            self.kpis[label].configure(text=str(int((analysis["Priority"] == label).sum())))
        manual = int((analysis["Data-chain judgement"] == "Needs Manual Review").sum())
        self.kpis["需人工确认"].configure(text=str(manual))

    def copy_tree_selection(self, tree):
        selected = tree.selection()
        if not selected:
            self.status.set("请先选择要复制的行")
            return "break"
        columns = list(tree["columns"])
        lines = ["\t".join(columns)]
        for item in selected:
            values = [cell_text(value) for value in tree.item(item, "values")]
            lines.append("\t".join(values))
        self.clipboard_clear()
        self.clipboard_append("\n".join(lines))
        self.status.set(f"已复制 {len(selected)} 行，可直接粘贴到 Excel")
        return "break"

    def _show_dataframe(self, tree, df, hidden_columns=None):
        if isinstance(tree, SpreadsheetGrid):
            tree.set_dataframe(df if df is not None else pd.DataFrame(), hidden_columns=hidden_columns)
            return
        tree.delete(*tree.get_children())
        if df is None or df.empty:
            tree["columns"] = []
            return
        display = df.copy()
        if hidden_columns:
            display = display[[col for col in display.columns if col not in hidden_columns]]
        if len(display) > 500:
            display = display.head(500)
        columns = [str(col) for col in display.columns]
        tree["columns"] = columns
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=self._smart_column_width(display, col), minwidth=56, stretch=False, anchor="w")
        for _, row in display.iterrows():
            values = [cell_text(row[col])[:600] for col in display.columns]
            tree.insert("", END, values=values)

    def _smart_column_width(self, df, col):
        name = str(col)
        lower = name.lower()
        long_text_cols = {
            "actual situation after check",
            "evidence-based reason",
            "process issues identified",
            "how to optimise",
        }
        if lower in long_text_cols:
            if lower in {"process issues identified", "how to optimise"}:
                return 520
            return 420
        if lower in {"chassis", "chassis num", "original_chassis", "normalized_chassis"}:
            return 180
        if lower in {"mismatch_type", "stock_type", "lifecycle_stage", "data-chain judgement"}:
            return 190
        if lower in {"priority", "confidence"}:
            return 95
        if lower in {"action owner", "owner", "status"}:
            return 120
        if any(token in lower for token in ["date", "pgi", "po_", "po ", "so_", "invoice", "bill", "matnr", "amount"]):
            return 140
        samples = [name] + [cell_text(value) for value in df[col].head(80)]
        max_len = max((len(value) for value in samples), default=10)
        return max(72, min(240, max_len * 9 + 24))


def main():
    try:
        app = App()
        app.mainloop()
    except ImportError as exc:
        print("缺少依赖：请安装 pandas 和 openpyxl 后再运行。", file=sys.stderr)
        print(exc, file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
